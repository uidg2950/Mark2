#=======================================================================#
# Binaries ELF Dependency Evaluator & Database Generator Tool           #
# Description: This present programm analyses a given set of files      #
# specified in a folder (and its subfolders) and retrieves all theirs   #
# relations by fetching ELF dependencies for each binary or (shared)    #
# library application files.                                            #
# Then the programm centralize all data into a json string and store it #
# into a text file, also adding time and uuid stamps, s othat unicity   #
# each analysis will be ensured.                                        #
#-----------------------------------------------------------------------#
# Language: generic Python Script. Tested on Conti Ubuntu.              #
# Author: Dr. Eugen Victor Cuteanu                                      #
# Department: A AN PL2 RD EMEA RBG SW APP2                              #
# Version: 0.1 draft / prototype.                                       #
# Date & Time: 12-03-2024 , 07:00am Germany.                            #
# Version: 0.2 draft / prototype, adding Excell Reporting               #
# Date & Time: 06-05-2024 , 08:00am Germany.                            #
# Version: 0.3 draft / prototype, adding csv output format              #
# Date & Time: 06-05-2024 , 08:00am Germany.                            #
# Version: 0.4 draft / prototype, linkdirs/owndirs extension.           #
# Date & Time: 23-05-2024 , 17:00am Germany.                            #
# Copyright: Continental Automotive AG                                  #
#=======================================================================#



#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~#
# Usage Examples:                                                                                  #
#    having default user current directory:  /home/uid41076  (or any uidABCDE own account folder)  #
#    with a complete workarea, like: ./drt16_repo_NAD21_drt15-sa515m-le21_16.y                     #
#    and python program located on home folder (so outside WA)                                     #
#                                                                                                  #
# Execute:                                                                                         #
#    python3 ./file_elfscan.py -f "./drt16_repo_NAD21_drt15-sa515m-le21_16.y/release/fs/" -t yes   #
#                                                                                                  #
#    python3 ./file_elfscan.py -f "./drt16_repo_NAD21_drt15-sa515m-le21_16.y/.build/prodfs" -t yes #
#        -l ./drt16_repo_NAD21_drt15-sa515m-le21_16.y/layers/project-drt15/drt15-sa515m-le21/      #
#        docs/FOSS-Compliance-Documentation/FOSS-Compliance-Documentation-INTERNAL -d 1            #
#        -c "./folders.json" -m 2                                                                  #
#                                                                                                  #
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~#

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from datetime import date
from datetime import datetime
import time
import uuid
import random
from os import listdir
from pathlib import Path
import os.path
import os
import base64
import platform
import io
import sys
import subprocess
import magic
import string
import json
import re
import csv
import argparse

#~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~#

""" provides status regarding existence of specified file """
def check_file_available(flnam):
	# check that file is really available:
	if ( (os.path.isfile(flnam) == True and os.path.exists(flnam) == True) ):
		return True;
	pass;
	return False;


""" function for removing an ending sub-string, mainly to avoind unnecessary separators. """
def remove_substring_from_end(original_string, substring):
    if original_string[-len(substring):] == substring:
        return original_string[:-len(substring)]
    else:
        return original_string

def string_into_array(xname):
	# Encapsulate string into one element arrray ['word.so\n']
	return xname.rsplit();

def wipe_so_sufix_out(flname):
	# Cut out the sufix after ".so" file extension!
	return (flname.split(".so", 1)[0]  + ".so");

def get_subcomponent_name(thepath):
	# Extract sub-component name, as being the substring after last slash!
	return (thepath.rsplit('/', 1)[1]);

def eliminate_white_chrs(strnam):
	# Wipe return characters out of the string.
	return strnam.strip('\n');

#~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~#


""" launch a command. For linux launch command(s) in a shell """
def launch_command(cmd, display=True, separator=' '):
    # show command's text !    
    if display:
        print(cmd)
    # This is done in order to have compatibility cross-platform
    # for windows to be able to use the command the shell has to be set to
    # false and true for Unix
    if os.name == "nt":
        use_shell = False
    else:
        use_shell = True
    # execute the command !
    processOutput = subprocess.Popen(cmd, stdin=None, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=use_shell)
    out, err = processOutput.communicate()
    #print( "Returncode is {}".format(processOutput.returncode) )
    if ( processOutput.returncode != 0  ):
        if display == True:
            print("ERROR at command >{}<".format(cmd))
            print(err)
        pass;
        output = err
        status = False
    else:
	# need conversion from stream to string.
        output = out.decode('utf-8').replace('\n',''+separator);
        status = True
    pass;
    # in case positive display argument, we will print the command's string!
    output = remove_substring_from_end(output,separator);
    if display == True:
        if len(output) > 0:
            print(output)
        else:
            print('')
    # return the status and output resulted after execution!
    return status, output 


""" obtain the mime signature of the specified file """
def get_file_info(file_path):
    try:
        mime = magic.Magic()
        file_info = mime.from_file(file_path)
    except Exception as e:
        return f'Error determining file type: {str(e)}'
    return file_info;
    

""" fetch the file type from the system using a dedicated coomand for this purpose """
def check_file_type(file_path):
    try:
        [status, output ] = launch_command("file --mime-type --brief " + file_path, False);
        mime_type = output.strip();
        # Check for specific keywords in the detailed output
        if 'executable' in mime_type or 'application' in mime_type:
            return 'Binary Executable'
        elif 'shared library' in mime_type or 'sharedlib' in mime_type:
            return 'Shared Library'
        elif 'archive' in mime_type or 'library' in mime_type:
            return 'Static Library'
        else:
            return 'Unknown'
    except subprocess.CalledProcessError:
        return 'Error determining file type'


""" retrieve a list with all binary files of interests """
def scan_files(spath, deepness=True, extens=""):
    # internal imbricated function
    def check_criteria(file_path,extens):
        rflag = False;
        # extra condition since only binaries are of interests for us!
        if ( extens == "" ):
            if ( check_file_type(file_path) != "Unknown" ):
                rflag = True;
            pass;
        # the other condition is that file name ends with specified substring!
        else:
            if ( file_path[-len(extens):] == extens ):
                rflag = True;
            pass;
        pass;
        return rflag;
    # the variable for files names list.
    file_list = [];
    # within exeception handling statament:
    try:
        if (deepness == True ):
            # get all files within subdirectories.
            k = 0;
            for root, dirs, files in os.walk(spath):
                for xfile in files:
                    file_path = os.path.join(root, xfile);
                    # debug counter!
                    if ( len(extens) > 0 and file_path.endswith(extens) ):
                        k=k+1;
                    # in case of met conditions, do accumulation
                    if ( check_criteria(file_path,extens) == True ):
                        file_list.append(file_path);
                    pass;
            if ( len(extens) > 0 ) :
                print("Found entries number: "+str(k));
            pass;
        else:
            # get only referece path within files.
            files = [os.path.join(spath, f) for f in os.listdir(spath) if os.path.isfile(os.path.join(spath, f))]
            file_list = [ fl for fl in files if check_criteria(fl,extens) == True ]
        pass;
    # in case of a an error:
    except Exception as e:
        print(f"Error scanning files: {str(e)}")
    pass;
    # return the list;
    return file_list;

#~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~#

""" This is a helper function for fetching the user's arguments! """
def parseargs():
    # create an instance for the argument parser.
    parser = argparse.ArgumentParser(description="automation of binaries dependency scanning based on readelf sofware tool.");
    # define possbile arguments.
    parser.add_argument('-f', action='store', dest='binfolder', help="Path to binaries representing deliverable software.", default='./');
    parser.add_argument('-l', action='store', dest='licfolder', help="Path to FOSS documentation with license about files.", default='NA');
    parser.add_argument('-w', action='store', dest='wrkfolder', help="Path to Workarea's sub-folder containing source code files.", default='./package');
    parser.add_argument('-a', action='store', dest='abtpath', help="Path to file containing the license information.", default='./myabout');
    parser.add_argument('-p', action='store', dest='dbspath', help="Path to file containing the database with entire report.", default='./myreport');
    parser.add_argument('-c', action='store', dest='cmppath', help="Path to file containing the database related to components folders.", default='./myfolders');
    parser.add_argument('-t', action='store', dest='timemark', help="Time stamping for processing period.", default='no');
    parser.add_argument('-i', action='store', dest='help', help="Help on available commands", default="none");
    parser.add_argument('-m', action='store', dest='mode', help="Mode of inventary: 4-excel only, 3-special, 2-extended (implicite), 1-simply, 0-basic(synch only).", default=2);
    parser.add_argument('-d', action='store', dest='deepness', help="Folder scanning mode 1-deep , 0-flat", default=1);
    # scan after parameters!
    args=parser.parse_args()
    pass;
    # print the chosen arguments.
    print(args)
    # return the content.
    return args;


""" This is an auxiliary function for argument parsing and tuning. """
def interpret_user_param(args):
    # local variable initializations
    myfolder = "";
    lcfolder = "";
    wkfolder = "";
    dbpath = "";
    lcpath = "";
    cmpath = "";
    xstamp = "";
    ymode = -1;
    zdeep = -1;
    # HELP LISTING:
    if ( len(args.help) > 0 and args.help != 'none' ) :
        if ( args.help == "!" ) :
            print("Available options are: -f [BINFOLDER] , -l [LICFOLDER] , -w [WRKFOLDER] , -a [ABTPATH] , -p [DBRPATH] , -c [CMPPATH] , -t [no/yes] , -m [0/1/2] , -d [0/1] , -i [!/?]");
        elif ( args.help == "?" ) :
            print("The available options may have the following options:");
            print("    -f [BINFOLDER] >> specify the relative or absolute path the folder containing binaries.");
            print("    -l [LICFOLDER] >> specify the relative or absolute path the folder containing FOSS documentation.");
            print("    -w [WRKFOLDER] >> specify the relative or absolute path the folder containing source code.");
            print("    -a [ABTPATH] >> specify the path the ABOUT file, containing all license data, to be created.");
            print("    -p [DBRPATH] >> specify the path to file, representing the database, to be created.");
            print("    -p [CMPPATH] >> specify the path to file, whose content represent folders list.");
            print("    -t [no/yes] >> indicates if a time stamping of the reports shall be made!");
            print("    -m [0/1/2/3/4] >> specify processing mode either basic synch only, simple, complete standard (default), special or excel only .");
            print("    -d [0/1] >> specify folder scan mode as deep or flat aka no subfolders .");
            print("    -i [!/?] >> This help indication. `!` = only option listing, `?` = detailed description.");
        else:
            print("Invalid specification! Use either '-i !' or '-i ?' for getting help information.");
        pass;
        exit(0);
    else:
        # VARIABLE INITIALIZATIONS: 
        if ( int(args.mode) > 5 ) :
            print("Unsupported mode! Default value taken for use.");
            ymode = 2;
        else:
            ymode = int(args.mode);
        pass;
        if ( int(args.deepness) > 1 ) :
            print("Unsupported scan strategy! Default value taken for use.");
            zdeep = 1;
        else:
            zdeep = int(args.deepness);
        pass;
        crtdir = os.getcwd();  # curdir = Path.cwd();
        if ( os.path.isdir(args.binfolder) == True and os.path.exists(args.binfolder) == True ):
            myfolder = args.binfolder;
        else:
            print("Current directory as binaries folder, since invalid path given!");
            myfolder = "./";
        pass;
        if ( os.path.isdir(args.licfolder) == True and os.path.exists(args.licfolder) == True ):
            lcfolder = args.licfolder;
        else:
            print("Current directory as license folder, since invalid path given!");
            lcfolder = "./";
        pass;
        if ( os.path.isdir(args.wrkfolder) == True and os.path.exists(args.wrkfolder) == True ):
            wkfolder = args.wrkfolder;
        else:
            print("Current directory as source code folder, since invalid path given!");
            wkfolder = "./";
        pass;
        fpath = os.path.abspath(args.dbspath);
        if ( os.path.isdir(fpath) == True and os.path.exists(fpath) == True ):
            dbpath = args.dbspath;
        else:
            print("Using self defined report file name, since invalid path given!");
            dbpath = "./elf_report.json";
        pass;
        apath = os.path.abspath(args.abtpath);
        if ( os.path.isdir(apath) == True and os.path.exists(apath) == True ):
            lcpath = args.abtpath;
        else:
            print("Using self defined about file name, since invalid path given!");
            lcpath = "./about_report.json";
        pass;
        cpath = os.path.abspath(args.cmppath);
        if ( os.path.isfile(cpath) == True and os.path.exists(cpath) == True ):
            cmpath = args.cmppath;
        else:
            print("Marking no file available, since invalid path given!");
            # cmpath = "./folders_report.json";
            cmpath = "n/a";
        pass;
        pass;
        if ( args.timemark != "no" and args.timemark != "yes" ) :
            xstamp = "unknown";
        elif ( args.timemark == "yes" ) :
            xstamp = "yes";
        elif ( args.timemark == "no" ) :
            xstamp = "no";
        else:
            xstamp = "n/a";
        pass;
    pass;
    # DEBUG informal prints:
    print("Scanning: " + myfolder + " deepness: " + str(zdeep) + " in mode: " + str(ymode) + " with stamping: " + xstamp + " !\n");
    print("Considering >> License Path: " + lcfolder + " & About file path: " + lcpath + " & Report file path: " + dbpath + " & Folders file path: " + cmpath + " !\n");
    # data multiplex delivered:
    return [myfolder , lcfolder , wkfolder, lcpath , dbpath , cmpath , ymode , zdeep , xstamp]; 


""" using time stamp and unique identifier, a JSON formated stamp will eb created """
def generate_uniquesignature(stamp):
    uniquestamp = "";
    # generating JSON footer, also check which stamping mode is requested.
    if ( stamp == "yes" ) :    
        uniquestamp = '"stamp" : { "date" : "' + str(datetime.now().strftime("%d-%m-%Y")) + \
                       '" , "time" : "' + str(datetime.now().strftime("%H:%M:%S")) + '" , "uuid" : "' + str(uuid.uuid4()) + '" }';
    elif ( stamp == "no" ) :    
        uniquestamp = '"stamp" : { "date" : "' + 'N/A' + \
                       '" , "time" : "' + 'N/A' + '" , "uuid" : "' + 'N/A' + '" }';
    else:
        uniquestamp = '"stamp" : { "date" : "' + str(random.randint(1, 1e9)) + \
                       '" , "time" : "' + str(random.randint(1, 1e9)) + '" , "uuid" : "' + str(random.randint(1, 1e9)) + '" }';
    pass;
    return uniquestamp;


""" function for making strings compatible with the json format """
def substitute_letters(string):
    substitutions = {'\\': '#', '/': '~', '"': '*', '\b': ' ', '\f': ' ', '\n': ' ', '\r': ' ', '\t': ' '}
    for key, value in substitutions.items():
        string = string.replace(key, value)
    return string

""" remove incompatible characters from a string as json format requires """
def substitute_characters(string):
    substitutions = {'\\': '#', '/': '~', '"': '*', '\b': ' ', '\f': ' ', '\n': ' ', '\r': ' ', '\t': ' '}
    translation_table = str.maketrans(substitutions)
    return string.translate(translation_table)


""" loads the various informations fields from about files out of specified folder """
def fetch_license_information(myfolder,stamp,aboutfile):
    file_list = scan_files(myfolder,True,".ABOUT");
    tags = [ "name" , "version" , "description" , "package_url" , "redistribute" , "modified" , "comments" , "version_full" , \
             "foss_distro_path" , "most_restrictive_license_category" , "package_origin" , "spdx_license_key_primary" ];
    # defining JSON strings:
    jsonhead_str = '{ "name": "Victorian ABOUT Database." , "version": "0.1" , "owner": "FOSS" , "components": [ ';
    # initialize local variables
    cnt = 0;
    knt = 0;
    record = "";
    entry = "";
    jsondbstr = jsonhead_str;
    # go through the list of files:
    for afilename in file_list:
        with open(afilename, 'r') as afilehndl:
            # information from new file to be recorded:
            record = "";
            cnt += 1;
            # starting with the seccond record we need to pre-append a continuation delimiter character,
            if ( cnt > 1 ):
                record += ', ';
            pass;
            # here we define the header of the new record:
            record += '{ "file" : ' + '"' + afilename + '" , "tags" : { ';
            knt = 0;
            # for each pair of {key,value} read from the file, do:
            for line in afilehndl:
                # condition for accepting a line from the ABOUT file:
                if any(tag == line[:len(tag)] for tag in tags):
                    # do the split between string representing key and the one of value.
                    aline = line.split(": ");
                    # we bypass in case of inconsistency aka invalid pair.
                    if ( len(aline) != 2 ) :
                        continue;
                    [ label , value ] = aline;
                    knt += 1;
                    # append a prolongcation delimiter starting with the seccond pair
                    if ( knt > 1 ) :
                        record += ', ';
                    pass;
                    # composing a pair entry for the JSON: remark, that double quotes character msut be removed from strings for the sake
                    # of JSON's standard, to avoid syntax errors !
                    entry = '"' + label.replace('"','~') + '" : "' + value.replace('"','~') + '"';
                    entry = entry.replace("\n","");
                    # accumulate to the present record associated with current file
                    record += entry;
                pass;
            pass;
            # adding local footer aka closure.
            record += ' } }';
        pass;
        # accumulating the record to the global JSON string.
        jsondbstr += record;
    pass;
    # generating JSON footer, also check which stamping mode is requested.
    jsonfoot_str = ' ] , ' + generate_uniquesignature(stamp) + ' }';
    # accumulate to main JSON string.
    jsondbstr += jsonfoot_str;
    # create JSON structure
    parsed_data = json.loads(jsondbstr);
    # Writing the Python dictionary to the JSON file
    if ( aboutfile != "" ) :
        with open(aboutfile, 'w') as json_file:
            json.dump(parsed_data, json_file, indent=4)
    pass;
    # deliver resultes data structure together with the main counter
    return [parsed_data , cnt];


""" perform the effective processing of the folder, considering all binary executable files """
def fetch_files_dependencies(myfolder, deepness, mode, stamp, dbreportfile):
    file_list = scan_files(myfolder,deepness);
    cnt = 0;
    jsondbstr = "";
    # defining JSON header:
    jsonhead_str = '{ "name": "Victorian ELF Database." , "version": "0.1" , "owner": "FOSS" , "files": [ ';
    # append to overall JSON string.
    jsondbstr += jsonhead_str;
    # formaly initialize local variables
    pnr = -1;
    pres_arstr = "[]";
    pres_folder = "[]";
    # for each file detected, do ELF dependency processing.
    for qfile in file_list:
        cnt += 1;
        finfo = get_file_info(qfile);
        # DEBUG: print("Considering the file: " + qfile + " => " + finfo + " ");
        # Define command string for fetching binary dependencies.
        scn_string = "readelf -d " + qfile + " | grep NEEDED ";
        # Executing the effective command:
        [answer , reply] = launch_command(scn_string, False, ',\n');
        # DEBUG: print(reply);
        # Extract fila names out of right brackets characters:
        if ( cnt > 1 ) :
            jsondelim_str = ' , ';
            jsondbstr += jsondelim_str;
        pass;
        # check dependency data availability.
        if ( answer == False ) :
            # DEBUG: print("No entry!");
            # define file entry in JSON string
            # jsonfile_str = '{ "path" : "' + qfile + '" ,  "mime" : "' + finfo.replace('"','~').replace('\\','#') + '" , "dependencies" : ' + '[]' + ' , "found" : ' + '[]' + ' , "linkdirs" : ' + '[]' + ' }';
            # using special function for adapting string:
            jsonfile_str = '{ "path" : "' + qfile + '" ,  "mime" : "' + substitute_characters(finfo) + '" , "dependencies" : ' + '[]' + ' , "found" : ' + '[]' + ' , "linkdirs" : ' + '[]' + ' }';
            # accumulate to global string
            jsondbstr += jsonfile_str;
        else:
            # extract relevant information, provided in rectangular brackets by the command.
            depflnams = re.findall(r'\[(.*?)\]',reply);
            # DEBUG: print("Dependecy array: " + str(depflnams));
            # reinitialize variables:
            pnr = 0;
            pres_arstr = "[ ";
            pres_folder = "[ ";
            # each dependency as file name will be processed in terms of availability
            for dfile in depflnams:
                # define string for file retrival command.
                src_string = "find " + myfolder + " -name " + dfile ;
                # execute the efective command
                [answer , response] = launch_command(src_string, False, ' | ');
                # DEBUG: print("We got: " + str(answer) + " with content: " + str(response) + " !");
                # small test to check if separator substring must be added.
                if ( pnr >= 1 ):
                    pres_arstr += " , ";
                    pres_folder += " , ";
                pass;
                # in case of non void response:
                if ( response != "" ):
                    # add information to local strings variables.
                    pres_arstr += '\"yes\"'
                    pres_folder += "\"" + response + "\"";
                else:
                    # otherwise add alternative substrings.
                    pres_arstr += '\"no\"'
                    pres_folder += '\"na\"'
                pass;
                # count the dependency file currently processed,
                pnr += 1;
            pass;
            # add local terminator character, required by JSON format.
            pres_arstr += " ]";
            pres_folder += " ]";
            # DEBUG: print("Dependency information arrays: " + str(pres_arstr) + " --- " + str(pres_folder) + " !\n");
            # define file information JSON substring:
            jsonfile_str = '{ "path" : "' + qfile +  '" , "mime" : "' +  substitute_characters(finfo) + \
                                                     '" , "dependencies" : ' + str(depflnams).replace("'","\"") + \
                                                      ' , "found" : ' + str(pres_arstr).replace("'","\"") + \
                                                      ' , "linkdirs" : ' + str(pres_folder) + ' }';
            # append it to the related global variable.
            jsondbstr += jsonfile_str;
        pass;
    pass;
    # generating JSON footer, also check which stamping mode is requested.
    jsonfoot_str = '] , ' + generate_uniquesignature(stamp) + ' }';
    # accumulate to main JSON string.
    jsondbstr += jsonfoot_str;
    # Convert JSON string to Python dictionary
    parsed_data = json.loads(jsondbstr);
    # DEBUG: print("Formateg JSON string is: \n " + str(parsed_data));
    # Writing the Python dictionary to the JSON file
    if ( dbreportfile != "" ) :
        with open(dbreportfile, 'w') as json_file:
            json.dump(parsed_data, json_file, indent=4)
    pass;
    # provide the data to the caller program.
    return [ parsed_data , cnt ] ;


""" this function scans the folder where the binaries are comming from """
def detect_origin_folder(depfile,srcfolder):
    # handling case of missing file exception!
    if ( check_file_available(depfile)==False ):
        print("Missing dependency file! Abort!");
        exit(2);
    pass;
    # define some variables for statistic assesment.
    nrtotal = 0;
    nrsteps = 20;
    nrcrt = 0;
    nrprev = 0;
    nrperc = 0;
    # using the database file in append rw+ mode!
    with open (depfile, 'r+') as elffile :
        print("Start processing file of " + str(os.path.getsize(depfile)) + " bytes ....");
        depdata = json.load(elffile);
        nrtotal = len(depdata["files"]);
        # iteration
        for fileentry in depdata["files"] :
            # DEBUG msg:
            #if ( "libsystemd" in fileentry["path"] ):
            #    print("scanning for: " + str(fileentry) + " .");
            #else:
            #    continue;
            #pass;

            nrcrt = nrcrt + 1;
            if ( nrcrt-nrprev >= (nrtotal/nrsteps-1) ) :
                nrperc = nrcrt * 100 / nrtotal;
                print("Processing item #"+str(nrcrt)+" , status: "+str(nrperc)+" % .");
                nrprev = nrcrt;
            pass;
            # get binary file name.
            elfpath = fileentry["path"];
            binname = elfpath.rsplit('/', 1)[1];
            scn_string = "find " + srcfolder + "  -type f  -name " + binname ;
            # Executing the effective command:
            [answer , reply] = launch_command(scn_string, False, ',\n');
            if ( reply != "" ):
                # DEBUG: print("Located @ \n"+reply + "\n - - - - - - - ");
                fileentry["located"] = str(reply);
            pass;

            # DEBUG code:
            #if ( "systemd" in fileentry["path"] ):
            #    exit(1);
            #pass;
        pass;
        print("Process completed 100% for " + str(nrtotal) + " entries !")
        # erase file content
        elffile.seek(0)  # Move the file pointer to the beginning 
        elffile.truncate()  # Clear the file contents for writing
        # write updated content from RAM memory to physical file on harddrive.
        json.dump(depdata, elffile, indent=4)
    pass;

#~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~#

""" effectuate database enhancements for ELF dependency json file based on ABOUT license and component directory json files. """
def synchronize_databases(licfile, depfile, dirfile):
    # open all three required files!
    if ( check_file_available(licfile)==True and check_file_available(depfile)==True and check_file_available(dirfile)==True ) :
        print("Files available for synch: " + licfile + " , " + depfile + " , " + dirfile + " !\n");
    else:
        print("Missing files for synchronization of databases! Abort!");
        exit(1);
    pass;
    with open(licfile, 'r') as abtfile, open(depfile, 'r+') as elffile, open(dirfile, 'r') as folfile:
        # load theirs content as josn databases.
        licdata = json.load(abtfile);
        depdata = json.load(elffile);
        dirdata = json.load(folfile);
        # go through the ELF dependency json database
        for fileentry in depdata["files"] :
            elfpath = fileentry["path"];
            # get binary file anme out of path string.
            binname = elfpath.rsplit('/', 1)[1];
            # local variable initializations
            cmp_name = "";
            cmp_base = "";
            cmp_paths = "";
            cmp_lic = "";
            cmp_most = "";
            cmp_spdx = "";
            flag = 0;
            # go through the folder dependency json database
            for specentry in dirdata["components"] :
                # fetch executable fila names separately
                specexecs = specentry["execs"];
                execsarray = specexecs.split(" ");  # SEPARATOR !?
                # go for each executable file name
                for bfname in execsarray :
                    # once we have a match ELF-Folder JSONs
                    if ( bfname == binname ) :
                       # fetch fields
                       cmp_name = specentry["name"];
                       cmp_base = specentry["base"];
                       specpaths = specentry["paths"];
                       # TODO: no need for separate path string storage!
                       # pathsarray = specpaths.split(";");
                       # cmp_paths = pathsarray.copy();
                       #
                       # copy paths strings !
                       cmp_paths = specpaths;
                       flag = 1;
                       break;
                    pass;
                pass;
                # once a match found, we can leave the loop.
                if ( flag == 1 ) :
                    break;
                pass;
            pass;
            # update ELF json with paths strings!
            fileentry["owndirs"] = cmp_paths;
            #
            flag = 0;
            cmp_label = "";
            # go through the ABOUT json database
            for compentry in licdata["components"] :
                # fetch fields
                licinfo = compentry["tags"];
                strname = licinfo["name"];
                # once we jave a match ABOUT-Folder JSONs
                if ( strname == cmp_name ) :
                    # fetch fields
                    prov_name = licinfo["package_origin"];
                    prov_lice = licinfo["version_full"];
                    prov_most = licinfo["most_restrictive_license_category"];
                    prov_spdx = licinfo["spdx_license_key_primary"];
                    # copy license information !
                    if ( prov_name != "Continental AG" ):
                        cmp_lic = prov_lice;
                        cmp_most = prov_most;
                        cmp_spdx = prov_spdx;
                    else:
                        cmp_lic = "proprietary";
                        cmp_most = "conti";
                        cmp_spdx = "n/a";
                    pass;
                    flag = 1;
                    cmp_label = cmp_name;
                    break;
                pass;
            pass;
            if ( flag == 0 ) :
                # repeat the above procedure of searching in ABOUT files by base name!
                for compentry in licdata["components"] :
                    licinfo = compentry["tags"];
                    strname = licinfo["name"];
                    if ( strname == cmp_base ) :
                        prov_name = licinfo["package_origin"];
                        prov_lice = licinfo["version_full"];
                        prov_most = licinfo["most_restrictive_license_category"];
                        prov_spdx = licinfo["spdx_license_key_primary"];
                        # check particular case of proprietary files.
                        if ( prov_name != "Continental AG" ):
                            cmp_lic = prov_lice;
                            cmp_most = prov_most;
                            cmp_spdx = prov_spdx;
                        else:
                            cmp_lic = "proprietary";
                            cmp_most = "conti";
                            cmp_spdx = "n/a";
                        pass;
                        flag = 1;
                        cmp_label = cmp_base;
                        break;
                    pass;
                pass;
            pass;
            # update ELF json with license tags!
            fileentry["license"] = cmp_lic;
            fileentry["most_restrict"] = cmp_most;
            fileentry["spdx_key"] = cmp_spdx;
            fileentry["label"] = cmp_label;
            fileentry["name"] = cmp_name; # record the name as used yt compile tiem as well !
        pass;
        # erase file content
        elffile.seek(0)  # Move the file pointer to the beginning 
        elffile.truncate()  # Clear the file contents for writing
        # write updated content from RAM memory to physical file on harddrive.
        json.dump(depdata, elffile, indent=4)
    pass;


""" function to detect the relation between file name and associated software link label! """
def scan_foreign_binaries(myfolder,depfile,othersdb):
    # create array:
    own_binaries = [];
    thr_binaries = [];
    # define strings:
    jsonhead_str = '{ "name": "Victorian OTHERS Database." , "version": "0.1" , "owner": "FOSS" , "files": [ ';
    jsondbstr = jsonhead_str;
    # open database
    with open(depfile, 'r+') as elffile:
        # load theirs content as josn databases.
        depdata = json.load(elffile);
        for fileentry in depdata["files"] :
            elfpath = fileentry["path"];
            binname = elfpath.rsplit('/', 1)[1];
            own_binaries.append(binname);
        pass;
        for fileentry in depdata["files"] :
            depbins = fileentry["dependencies"];
            for depfile in depbins:
                if ( not depfile in own_binaries ) :
                    thr_binaries.append(depfile);
                    break;
                pass;
            pass;
        pass;
    pass;
    unique_frgn = list(dict.fromkeys(thr_binaries));
    # DEBUG:
    # print(unique_frgn);
    kfe = 0;
    label = "";
    for elem in unique_frgn :
        # define command for file search
        src_string = "find " + myfolder + " -name " + elem ;
        # execute the efective command
        [answer , response] = launch_command(src_string, False, ' , ');
        oth_path=response.split(" , ")[0];
        # check for response availability:
        if ( len(oth_path) > 0 ) :
            # define and execute command for information retrieval including symbolic link
            [status, output ] = launch_command("ls "+oth_path+" -l", False);
            sym_lnk = output.strip().split(" -> ")[1];
        else:
            sym_lnk = "none";
        pass;
        # check for responsa avilable
        if ( not sym_lnk == "none" ) :
            # check if new name is in our json
            if ( not sym_lnk in own_binaries ) :
                label = "no";
            else:
                label = "yes";
            pass;
        else:
            label = "none";
        pass;
        # handling separator string
        if ( kfe > 0 ):
            jsondbstr += ', '
        else:
            kfe+=1;
        pass;
        # compose and accumulate json entry!
        jsondbstr += '{ "file" : "'+str(elem)+'" , "symbol": "'+str(sym_lnk)+'" , "own" : "'+str(label)+'" }';
    pass;
    # define json string
    jsonfoot_str = '] , ' + generate_uniquesignature("yes") + ' }';
    jsondbstr += jsonfoot_str;
    # DEBUG:
    # print(jsondbstr);
    parsed_data = json.loads(jsondbstr);
    # in case of specified filename, then save the json string.
    if ( othersdb != "" ) :
        with open(othersdb, 'w') as json_file:
            json.dump(parsed_data, json_file, indent=4)
    pass;


""" function to merge the dependencies array contents of entries having the same component name """
def merge_json_records(data):
    # Create a copy of the input data to preserve other keys
    result_data = {key: value for key, value in data.items() if key != 'relations'}
    result_data['relations'] = []
    # local data array
    merged_records = []
    # go through all records in the json data
    for record in data['relations']:
        component = record['component']
        dependencies = set(record['dependencies'])  # Use a set to handle duplicates
        # initially assuming no match
        found = False
        # go through all currently existing collected entries
        for merged_record in merged_records:
            # once same component name string found
            if merged_record['component'] == component:
                # Merge dependencies
                merged_record['dependencies'].update(dependencies)
                found = True
                break
        pass;
        # if nothing found add all at once
        if not found:
            # Add new entry
            merged_records.append({
                'component': component,
                'dependencies': dependencies
            })
        pass;
    pass;
    # Convert sets back to lists for the final output
    for record in merged_records:
        result_data['relations'].append({
            'component': record['component'],
            'dependencies': list(record['dependencies'])
        })
    pass;
    # Get the number of entries in the modified records array
    records_count = len(result_data['relations'])
    print("Newly merged date do contain " + str(records_count) + " entries int the relations array.");
    # deliver data structure
    return result_data


""" funtion for generating component level dependency relations """
def analyse_dependencies(depfile,anlfile):
    # Initialize an empty dictionary
    my_dict = {}
    # Function to add an element to the dictionary if the key does not exist, and return the value
    def add_or_get_value(key, value):
        if key not in my_dict:
            my_dict[key] = value
            #print(f'Added key: {key}, value: {value}')
            return ""
        else:
            #print(f'Key {key} already exists with value: {my_dict[key]}')
            return my_dict[key]
        pass;
    pass;
    # Function to add a value to the list associated with the key, or create a new list if the key does not exist
    def add_or_append_value(key, value):
        if ( value == "" ) :
            return "";
        if key not in my_dict:
            my_dict[key] = value;
            #print(f'Added key: {key}, value: [{value}]')
        else:
            my_dict[key].append(value)
            #print(f'Appended value: {value} to key: {key}, new list: {my_dict[key]}')
        return my_dict[key]
    pass;
    # operates only when db file available!
    if ( check_file_available(depfile)==False ):
        print("Unavailable dependency file! Abort!");
        exit(1);
    pass;
    #
    cnt = 0;
    jsondbstr = "";
    # defining JSON header:
    jsonhead_str = '{ "name": "Victorian Dependency Database." , "version": "0.1" , "owner": "FOSS" , "relations": [ ';
    # append to overall JSON string.
    jsondbstr += jsonhead_str;
    # empy the dictionary
    my_dict.clear();
    # start the work
    with open(depfile, 'r') as elffile:
        depdata = json.load(elffile);
        allentries = depdata["files"] ;
        # go through the ELF dependency json database
        for fileentry in allentries:
            elfpath = fileentry["path"];
            # get binary file anme out of path string.
            binname = elfpath.rsplit('/', 1)[1];
            cmpname = fileentry["name"];
            dependy = fileentry["dependencies"];
            depcomp = "";
            deplist = [];
            # skip no name entries
            if ( cmpname == "" ) :
                continue;
            # going through all dependency files
            for usedbin in dependy :
                # check if the binary where looked up after its component name.
                if ( usedbin not in my_dict ):
                    # mark strong name criteria:
                    namflag = False;
                    # go thriugh all records to find its associated component name.
                    for crtentry in allentries:
                        elfpstr = crtentry["path"];
                        filname = get_subcomponent_name(elfpstr)
                        if ( usedbin == filname ) :
                            namflag = True;
                            depcomp = crtentry["name"];
                            add_or_append_value(usedbin,depcomp);
                            break;
                        pass;
                    pass;
                    if ( not namflag ) :
                        for crtentry in allentries:
                            elfpstr = crtentry["path"];
                            filname = get_subcomponent_name(elfpstr)
                            if ( wipe_so_sufix_out(usedbin) == wipe_so_sufix_out(filname) ) :
                                namflag = True;
                                depcomp = crtentry["name"];
                                add_or_append_value(usedbin,depcomp);
                                break;
                            pass;
                        pass;
                    pass;
                else:
                    # here we know ehich is its component belonging to.
                    depcomp = my_dict[usedbin];
                pass;
                # insert component name to dependency list.
                if ( depcomp != "" ) :
                    deplist.append(depcomp);
            pass;
            uniquedeps = list(dict.fromkeys(deplist));
            jsonentry = '{ "component" : "' + cmpname + '" , "dependencies" : ' + str(uniquedeps).replace("'","\"") + ' }';
            # handling separator string
            if ( cnt > 0 ):
                jsondbstr += ', '
            else:
                cnt+=1;
            pass;
            jsondbstr += jsonentry;
        pass;
    pass;
    # generating JSON footer, also check which stamping mode is requested.
    jsonfoot_str = '] , ' + generate_uniquesignature("yes") + ' }';
    # accumulate to main JSON string.
    jsondbstr += jsonfoot_str;
    # Convert JSON string to Python dictionary
    parsed_data = json.loads(jsondbstr);
    # Ensure unique merged entries
    merged_data = merge_json_records(parsed_data)
    # DEBUG: print(merged_data);
    # in case of specified filename, then save the json string.
    if ( anlfile != "" ) :
        with open(anlfile, 'w') as json_file:
            json.dump(merged_data, json_file, indent=4)
    pass;


""" this functions re-format the dependency data and store them into an excell file """
def convert_elf_to_excel(xlsfile,depfile):
    # create and empty excel workbook into a new file.
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Hello"
    sheet2 = wb.create_sheet(title="FOSS")
    wb.save(xlsfile)
    # setup table header.
    workbook = load_workbook(xlsfile);
    sheet = workbook.active;
    sheet["A1"] = "Path"
    sheet["B1"] = "owndirs"
    sheet["C1"] = "located"
    sheet["D1"] = "Most Restrictive License Cath."
    sheet["E1"] = "Spdx License Key"
    sheet["F1"] = "Component Name"
    sheet["G1"] = "Dependencies"
    sheet["H1"] = "linkdirs"
    # prepare for extra facelift of the Excel sheet.
    # Color Indexes, ranges from 0-63
    colors = [6, 3, 48, 37, 28, 7, 58, 24, 35, 19]
    fillers = []
    for color in colors:
        temp = PatternFill(patternType='solid', fgColor=Color(indexed=color));
        fillers.append(temp)
    pass;
    # setting up color schem afor first row aka table's header.
    sheet["A1"].fill = fillers[1];
    sheet["B1"].fill = fillers[2];
    sheet["C1"].fill = fillers[3];
    sheet["D1"].fill = fillers[4];
    sheet["E1"].fill = fillers[5];
    sheet["F1"].fill = fillers[6];
    sheet["G1"].fill = fillers[7];
    sheet["H1"].fill = fillers[8];
    # start to fill the data into the cells.
    line_num = 2;
    with open(depfile, 'r+') as elffile:
        # load theirs content as josn databases.
        depdata = json.load(elffile);
        for fileentry in depdata["files"] :
            # fields retrieval_:
            elfpath = fileentry["path"];
            elfdeps = fileentry["dependencies"];
            elflink = fileentry["linkdirs"];
            elfownd = fileentry["owndirs"];
            if ( "located" in fileentry.keys() ):
                elflocd = fileentry["located"];
            else:
                elflocd = "n/a";
            pass;
            # elfdecl = fileentry["license"];
            elfmost = fileentry["most_restrict"];
            elfspdx = fileentry["spdx_key"];
            elfname = fileentry["name"];
            # populating data in separate line for each dependency sub-entry!
            scnt = 0;
            for depfile in elfdeps:
                sheet[f"A{line_num}"].value = elfpath;
                sheet[f"B{line_num}"].value = str(elfownd);  # compile-linking based detected folder path.
                sheet[f"C{line_num}"].value = str(elflocd);  # find command based scaned folder path.
                sheet[f"D{line_num}"].value = elfmost;
                sheet[f"E{line_num}"].value = elfspdx;
                sheet[f"F{line_num}"].value = elfname;
                sheet[f"G{line_num}"].value = depfile;
                sheet[f"H{line_num}"].value = str(elflink[scnt]);
                # next dependency index
                scnt = scnt + 1;
                # update current line
                line_num = line_num + 1
            pass;
        pass;
    pass;
    # Autofit the worksheet.
    # to SIMULATE: sheet.autofit()
    # Iterate over all columns and adjust their widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        pass;
        adjusted_width = (max_length + 2) * 1;
        sheet.column_dimensions[column_letter].width = adjusted_width;
    pass;
    # update the file !
    workbook.save(filename=xlsfile);
    # Open a CSV file for writing
    with open('foss.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        # Iterate over the rows in the worksheet
        for row in sheet.iter_rows(values_only=True):
            # Write each row to the CSV file
            writer.writerow(row)
        pass;
    # number of entries:
    return line_num;

#~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~#

""" this condition ensures that the following code will execute only by direct invokation of this module """
if __name__ == '__main__':
    # get input and output file.
    args = parseargs();
    # Get marker strings.
    stamp_start = str(date.today()) + " # " + str(datetime.now().time());
    signature = str(uuid.uuid4());
    # Show them.
    print("Start on = " + stamp_start + " \n" );
    print("Signature: " + signature + " \n" );
    # process the arguments.
    [myfolder , lcfolder , wkfolder , lcpath , dbpath , cmpath , ymode , zdeep , xstamp] = interpret_user_param(args);

    # we do data base creation only in case of non basic operation required.
    if ( ymode == 1 or ymode == 2 ) :
        # Preliminary collection routine!
        print("Refetching license information....\n");
        [ parsed_data , cnt ] = fetch_license_information(lcfolder,xstamp,lcpath);
        print("Generated JSON string: \n" + str(parsed_data) + "\n" + "~" * 80 + "\n");
        # Effective processing routine!
        print("Refetching binaries dependencies....\n");
        [ parsed_data , cnt ] = fetch_files_dependencies(myfolder,zdeep,ymode,xstamp,dbpath);
        print("Generated JSON string: \n" + str(parsed_data) + "\n" + "~" * 80 + "\n");
    pass;

    # we do synch only if basic operation or extensive mode are requested.
    if ( ymode == 0 or ymode == 2 ) :
        print("Syncing databases....\n");
        synchronize_databases(lcpath,dbpath,cmpath);
        scan_foreign_binaries(myfolder,dbpath,"othersdb.json");
        convert_elf_to_excel("foss.xlsx",dbpath);
    pass;

    if ( ymode == 3 ):
        print("Enhancing binaries database....\n");
        detect_origin_folder(dbpath,wkfolder);
        convert_elf_to_excel("foss.xlsx",dbpath);
    pass;

    if ( ymode == 4 ):
        convert_elf_to_excel("foss.xlsx",dbpath);
    pass;

    if ( ymode == 5 ):
        analyse_dependencies(dbpath,"./depends.json");
    pass;

    # Get yet another timer string.
    stamp_stop = str(datetime.now().strftime("%d-%m-%Y")) + " ~ " + datetime.now().strftime("%H:%M:%S");
    # show it on console.
    print("Stop on = " + stamp_stop + " \n" );
    # apply signature into report file, if requested.
    
    
#~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~#
""" EOF """
