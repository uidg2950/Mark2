import os
import re
import json
import datetime
import glob
import time
import urllib
from collections import OrderedDict

from openpyxl import load_workbook

from .get_jazz_data import *
from .get_artifactory_data import *
from .klocwork_findings import main as klocwork_findings_main

from .settings import *
from helper.excel_helper import check_sheetname_in_excel
from helper.credentials_keyring import get_password

# Jira Query
from cmlib.jira.query import Query

release_note = OrderedDict()

def get_jira_content(filter, get_no_of_findings=False):
    """
    The get_jira_content function returns a list of issues from the Jira API.
    The function takes two arguments:
        - filter: A string containing the jql query to be used in order to get the issues.
        - get_no_of_findings: A boolean indicating whether or not we want to return only
          the number of findings (True) or all fields for each finding (False). Default is False.

    :param filter: Specify the search criteria
    :param get_no_of_findings=False: Get the number of findings in a given filter
    :return: A list of issues or number of findings
    """
    max_results = 1
    while True:
        url = "{}{}&maxResults={}".format(jira["api_search_url"], filter, max_results)

        res = requests.get(url, headers=jira["header"])
        if res.status_code != 200:
            raise Exception("The status code of the jira api is not 200!")

        if max_results == res.json()["total"] or get_no_of_findings:
            break
        else:
            max_results = res.json()["total"]

    if get_no_of_findings:
        return res.json()["total"]
    else:
        return res.json()["issues"]


def get_list_of_changes_to_previous_release_from_jira(defects_jql_query):
    """
    The get_list_of_changes_to_previous_release_from_jira function takes a string as an argument.
    It then queries JIRA for all issues that have the given string in their labels field.
    The function then parses through each issue's summary and extracts the id of the previous release from it.
    This is done by searching for text enclosed in square brackets at the end of each issue's summary, which should be something like [#12345].
    The number between those brackets is extracted and added to a list, which is returned.

    :param defects_jql_query: Filter the issues from jira
    :return: A list of all the bugs that were fixed in the previous release
    """

    filter = "jql=type={}".format(defects_jql_query[5:].replace("&quot;", "'"))
    issues = get_jira_content(filter)

    result_list = []

    # filter the id of the previous release from the issue summary
    for issue in issues:
        summary = issue["fields"]["summary"]
        try:
            previous_release_number = summary.rsplit("[", 1)[1].split("]")[0]
            result_list.append(int(previous_release_number))
        except:
            pass

    # filter all double entries from list
    result_list = list(set(result_list))
    return result_list


def set_open_defects():
    """
    The set_open_defects function is used to get the number of open defects in JIRA.
    It does this by using a filter that looks for all tickets with the following criteria:
    - Open tickets -> Problem reports not closed
    - Tickets assigned -> Problem reports already assigned
    - tickets a/b/c -> Problem reports discovered during System Test with priorities Blocking/High, Medium and Low

    :return: The number of open defects
    """
    # Jira API
    jira_querier = Query()

    searches = {
                "open_tickets_total" : '"Requesting Projects" in ("CONMOD Customer Solution") AND issuetype in ("Problem Report (PR)") AND status != Closed',
                "tickets_assigned" : '"Requesting Projects" in ("CONMOD Customer Solution") AND issuetype in ("Problem Report (PR)" ) AND status != Closed  AND assignee is not EMPTY',
                "tickets_a" : '"Requesting Projects" in ("CONMOD Customer Solution") AND issuetype in ("Problem Report (PR)" ) AND status != Closed AND "Detected by" = "System Test" AND ( priority = Blocking OR priority = High)',
                "tickets_b" : '"Requesting Projects" in ("CONMOD Customer Solution") AND issuetype in ("Problem Report (PR)" ) AND status != Closed AND "Detected by" = "System Test" AND ( priority = Medium )',
                "tickets_c" : '"Requesting Projects" in ("CONMOD Customer Solution") AND issuetype in ("Problem Report (PR)" ) AND status != Closed AND "Detected by" = "System Test" AND (priority = Low )'
                }

    for ticket_type, query in searches.items():
        print("Looking for {}....".format(ticket_type))
        tickets_count = jira_querier.search_tickets(query,simple_search=True)
        tickets_count = len(tickets_count)
        release_note["kpis"].update({ ticket_type : tickets_count})

def get_kpis_from_excel(baseline_version, path=str, sheetname=str, key_value_cell_for_release_note=OrderedDict):
    """
    The get_kpis_from_excel function reads the latest excel file in a given directory and returns
    the values of the cells specified by key_value_cell_for_release_note. The function is called
    in main().

    :param path=str: Define the path to the excel file
    :param sheetname=str: Specify the sheetname in the excel file
    :param key_value_cell_for_release_note=OrderedDict: Define the cell in which the key and value are stored
    :return: A dictionary with the kpis from the excel file
    """
    if os.name == "posix":
        path_splitted = path.split("/")
        new_mounting_point = os.path.sep.join(path_splitted[0:-1])
        local_folder_for_mount = path_splitted[-2]
        mount_command = "sudo mount.cifs -o credentials=/home/uidg4627/.credentials/dpfil02 {} {}".format(new_mounting_point, local_folder_for_mount)
        print(mount_command)
        os.mkdir(local_folder_for_mount)
        os.system(mount_command)
        path = os.path.join(local_folder_for_mount, "*")

    list_of_files_tmp = glob.glob(path)
    # Filter files out, which have a "~" in name => temporary created
    list_of_files = []
    for element in list_of_files_tmp:
        if "~" not in element:
            list_of_files.append(element)

    # get newest file
    latest_file = max(list_of_files, key=os.path.getctime)

    # read csv
    print(latest_file)
    wb = load_workbook(latest_file)

    # Identifying correct column
    results_wb = wb[sheetname]

    # filters for clusters, "N/A" is applicable to both
    cluster_value = "CL_43" if "cl43" in baseline_version else "CL_45"
    row_values = [ cluster_value, "N/A"]
    rv = []
    print("Filters for cluster: {}".format(row_values))

    for value in row_values:
        for row in results_wb.rows:
            for t in row:
                if t.value == value:                                   # i.e. <Cell 'SyRD TC Conmod_Cluster Summary'.B3>
                    val_in_row = str(t)
                    val_in_row = re.sub(r"[\<\>]", "", val_in_row)     # i.e. SyRD TC Conmod_Cluster Summary.B3
                    rv.append(val_in_row.split(".")[1][1])             # row value = rv

    for key, column in key_value_cell_for_release_note.items():
        for row in results_wb.rows:                     # listing all rows
            for t in row:
                if t.value == column:
                    cv = str(t)                         # column value = cv
                    cv = re.sub(r"[\<\>]", "", cv)
                    cv = cv.split(".")[1][0]
                    key_value_cell_for_release_note[key] = (cv + rv[0], cv + rv[1])

    check_sheetname_in_excel(wb, sheetname, latest_file)

    sheet = wb[sheetname]
    for key, value in key_value_cell_for_release_note.items():
        total_value = int(sheet[value[0]].value + sheet[value[1]].value)
        release_note["kpis"].update({key: total_value})

    if os.name == "posix":
        time.sleep(3)
        os.system("sudo umount {}".format(local_folder_for_mount))
        time.sleep(3)
        os.rmdir(local_folder_for_mount)


def set_cpu_ram_rom_data(xlsx_file):
    """
    The set_cpu_ram_rom_data function :
        - gets the CPU idle and top command from RAM+CPU - Application Processor sheet in xlsx_file
        - gets the ressource_ram_* and ressource_rom_* from RAM-Split SoC Cores sheet in xlsx_file

    :return: None
    """
    wb = load_workbook(xlsx_file, data_only=True)

    # get ressource_cpu_measure_release
    sheetname = "RAM+CPU - Application Processor"
    check_sheetname_in_excel(wb, sheetname, xlsx_file)
    sheet = wb[sheetname]
    cpu_idle = sheet["A1"].value
    for x in range(4, 30): # This range doesn't obey any particular reason, so that, this can be modified under discretion
        if "top" in sheet[f"H{x}"].value:
            cpu_top_cmd_str = sheet[f"H{x}"].value
            cpu_top_cmd = sheet[f"G{x}"].value
            break
    else:
        raise Exception("There is not the 'top' command in the cells H4 - H8 of {} in {}".format(sheetname, xlsx_file))
    cpu_idle_list = cpu_idle.split()
    for i, element in enumerate(cpu_idle_list):
        if "idle" in element:
            cpu_idle_int = int(cpu_idle_list[i - 1].split("%")[0])
    cpu_eso = 0
    for x in range(4, 300):
        if sheet[f"G{x}"].value == "0%":
            break
        if "eso" in sheet[f"H{x}"].value:
            cpu_eso_value = int(sheet[f"G{x}"].value.split("%")[0])
            cpu_eso += cpu_eso_value
    cpu_top_cmd_int = int(cpu_top_cmd.split("%")[0])
    ressource_cpu_measure_release = int(100 - cpu_idle_int - cpu_top_cmd_int - cpu_eso)
    release_note["kpis"].update({"ressource_cpu_measure_release": "{}%".format(ressource_cpu_measure_release)})

    # get ressource_ram_*
    sheetname = "RAM - Split SoC Cores"
    check_sheetname_in_excel(wb, sheetname, xlsx_file)
    sheet = wb[sheetname]
    release_note["kpis"].update({"ressource_ram_applicable": "{:.2f} MB".format(float(sheet["B5"].value))})
    b_11 = float(sheet["B11"].value)
    b_12 = float(sheet["B5"].value) - b_11
    release_note["kpis"].update({"ressource_ram_limit_release": "{:.2f} MB".format(b_11 + b_12)})
    release_note["kpis"].update({"ressource_ram_measure_release": "{:.2f} MB".format(float(sheet["B19"].value))})

    # get ressource_rom_*
    sheetname = "Flash - NAND Partition Info"
    sheet = wb[sheetname]
    b_30 = sheet["B30"].value
    release_note["kpis"].update({"ressource_rom_applicable": "{:.2f} MB".format(float.fromhex(b_30) / (1000 * 1000))})
    release_note["kpis"].update({"ressource_rom_limit_release": "{:.2f} MB".format(float.fromhex(b_30) / (1000 * 1000))})

    sheetname = "Flash Usage - UBI0 Volume"
    check_sheetname_in_excel(wb, sheetname, xlsx_file)
    sheet = wb[sheetname]
    b_11 = 0
    for x in range(2, 10):
        b_11 += int(sheet["B{}".format(x)].value)
    b_9 = int(sheet["B9"].value)
    b_15 = b_11 - b_9
    e_15 = b_15 / 1000 / 1000
    release_note["kpis"].update({"ressource_rom_measure_release": "{:.2f} MB".format(float(e_15))})


def main(baseline_version, release_id, rel_note, test_plan_id, output_folder, defects_jql_query):
    """
    The main function of the script.

    :param rel_note: Get the version number and call week from it
    :param test_plan_id: Get the number of software tests faulty
    :param output_folder: Specify the output folder
    :param defects_jql_query: Get the list of changes to previous release from jira
    :return: The release note as a json file
    """
    version = rel_note.split("G_V")[1].split("_CW")[0]
    call_week = rel_note.split("CW")[1]
    call_week_full_naming = "CONMOD.CW{}".format(call_week)

    # add releasenote data
    release_note.update({"releasenote": OrderedDict()})
    release_note["releasenote"].update({"software_name_supplier": "ModemBSW (SWC / DK1)"})
    release_note["releasenote"].update({"software_part_number_vw": "887"})
    release_note["releasenote"].update({"software_name": "ModemBSW (SWC / DK1)"})
    sw_version = "conmod-sa515m-{}".format(version)
    release_note["releasenote"].update({"sw_version": sw_version})
    release_note["releasenote"].update({"number_release_note": rel_note})
    release_note["releasenote"].update({"version": sw_version})
    release_note["releasenote"].update({"date": datetime.date.today().strftime("%Y-%m-%d")})
    release_note["releasenote"].update({"list_changes_to_previous_release": get_list_of_changes_to_previous_release_from_jira(defects_jql_query)})
    release_note["releasenote"].update({"list_supplied_evidences": ["Release notes", "NAD Test Report", "Traceability Report"]})

    # add kpi data
    release_note.update({"kpis": OrderedDict()})

    # set software_requirements Part I
    # \\cw01.contiwan.com\SMT\didr2537\Reports\CONMOD\Traceability_reports
    # path = os.path.join(r'\\', "cw01.contiwan.com", "SMT", "didr2537", "Reports", "CONMOD", "Traceability_reports", "*")
    path = os.path.join(r'{}{}'.format(os.path.sep, os.path.sep), "automotive-wan.com", "SMT", "didr2537", "Reports", "CONMOD", "Traceability_reports", "*")
    sheetname = "SyRD TC Conmod_Cluster Summary"
    key_value = OrderedDict({
        "software_requirements": "Total",
        "software_requirements_test": "SyRD Requirements Covered by TC"
    })

    get_kpis_from_excel(baseline_version, path, sheetname, key_value)

    # set software_requirements Part II
    # path = os.path.join(r'\\', "cw01.contiwan.com", "SMT", "didr2537", "Reports", "CONMOD", "Testcase_Execution_Report", "*")
    path = os.path.join(r'{}{}'.format(os.path.sep, os.path.sep), "automotive-wan.com", "SMT", "didr2537", "Reports", "CONMOD", "Testcase_Execution_Report", "*")
    sheetname = "Test Result by Conmod_Cluster"
    key_value = OrderedDict({
        "software_requirements_fully_tested": "Passed"
    })

    get_kpis_from_excel(baseline_version, path, sheetname, key_value)

    if test_plan_id:
        nr_software_tests_faulty = get_number_of_software_tests_faulty(test_plan_id)
    else:
        nr_software_tests_faulty = 0
    release_note["kpis"].update({"software_tests_faulty": nr_software_tests_faulty})
    release_note["kpis"].update({"software_error_tickets": nr_software_tests_faulty})

    set_open_defects()

    release_note["kpis"].update({"ressource_cpu_applicable": "100%"})
    release_note["kpis"].update({"ressource_cpu_limit_release": "100%"})

    kpi_report_xlsx_file = os.path.join(output_folder, "kpi_report.xlsx")
    download_kpi_data_xlsx_file_from_artifactory(baseline_version, release_id, kpi_report_xlsx_file)
    # downlaod_kpi_data_xlsx_file(version, output_folder)
    set_cpu_ram_rom_data(kpi_report_xlsx_file)

    # Get Klocwork findings
    severity_results = klocwork_findings_main(sw_version.upper(), baseline_version, release_id)
    release_note["kpis"].update({"serious_difference": severity_results["1"]})
    release_note["kpis"].update({"considerable_difference": severity_results["2"]})
    release_note["kpis"].update({"marginal_difference": severity_results["3"]})
    release_note["kpis"].update({"no_or_slight_difference": "no figures collected"})

    # write result to .json file
    json_release_note = json.dumps(release_note, indent=2)
    print(json_release_note)
    output_file = os.path.join(output_folder, "SWC_Release_Notes_{}_Continental.json".format(version))
    with open(output_file, "w") as file:
        file.writelines(json_release_note)
